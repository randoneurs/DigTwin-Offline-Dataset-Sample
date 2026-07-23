#!/usr/bin/env python3
"""Convert a RETINA outlet-visit Excel export ("Retina - <Month> <Year> (Monthly).xlsx")
into the JSON snapshot index.html's RETINA card loads.

Usage:
  python3 scripts/build_retina_snapshot.py \
      --xlsx "Retina - June 2026 (Monthly).xlsx" \
      --month 2026-06 \
      --out retina_2026-06.json

Source columns (in the sheet named after the file, not "Sheet1" which is a stray
pivot table): Created Date, Store ID, Store Name, Store Area, Store Region,
Store Branch, Store City, Operator, Availability Score, Visibility Score,
AV Index, Quadrant. One row per outlet-visit-operator observation across the
month; there can be several visits per store in a month and not every store
carries every operator.

Aggregation: average every raw-brand observation (Availability, Visibility)
at each geography level (City, Branch, Region, Area, Nationwide) — a plain
mean, no weighting — then combine the 7 raw brands into the 3 tracked
operator groups (also a plain mean of the constituent brands' city-level
scores): telkomsel = {telkomsel, byu}, ioh = {indosat, tri}, xlsmart =
{xl, axis, smartfren}. AV Index is re-derived as (availability+visibility)/2
after averaging rather than averaging the source AV Index column directly —
equivalent for a plain mean, but keeps the two scores as the source of truth.
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

AREA_MAP = {
    "Area 1": ["SUMBAGUT", "SUMBAGTENG", "SUMBAGSEL"],
    "Area 2": ["JAKARTA-BANTEN", "EASTERN JABOTABEK", "JABAR"],
    "Area 3": ["JATENG-DIY", "JATIM", "BALINUSRA"],
    "Area 4": ["KALIMANTAN", "SULAWESI", "PUMA"],
}
REGION_TO_AREA = {region: area for area, regions in AREA_MAP.items() for region in regions}

# RETINA's "Store Region" column (any case/spacing) -> this dashboard's REGION_LABELS key.
REGION_XLSX_TO_KEY = {
    "sumbagut": "SUMBAGUT",
    "sumbagteng": "SUMBAGTENG",
    "sumbagsel": "SUMBAGSEL",
    "jakarta banten": "JAKARTA-BANTEN",
    "eastern jabotabek": "EASTERN JABOTABEK",
    "jabar": "JABAR",
    "jateng-diy": "JATENG-DIY",
    "jatim": "JATIM",
    "bali nusra": "BALINUSRA",
    "kalimantan": "KALIMANTAN",
    "sulawesi": "SULAWESI",
    "maluku dan papua": "PUMA",
}

# The dashboard's curated CITY_MAP (index.html) — only these cities resolve
# through the fArea/fRegion/fCity filters. RETINA's city spelling matches the
# same "KOTA X" / bare-kabupaten-name convention as the META market-share
# export, including the same handful of exceptions.
CITY_MAP = {
    "SUMBAGUT": ["Kota Medan", "Kota Banda Aceh", "Kabupaten Deli Serdang", "Kota Pematangsiantar", "Kota Binjai", "Kabupaten Langkat", "Kota Lhokseumawe", "Kabupaten Simalungun"],
    "SUMBAGTENG": ["Kota Pekanbaru", "Kota Padang", "Kota Batam", "Kota Jambi", "Kota Dumai", "Kabupaten Kampar", "Kota Tanjungpinang", "Kabupaten Bungo"],
    "SUMBAGSEL": ["Kota Palembang", "Kota Bandar Lampung", "Kota Bengkulu", "Kota Pangkalpinang", "Kabupaten Ogan Ilir", "Kota Prabumulih", "Kabupaten Lampung Selatan", "Kota Metro"],
    "JAKARTA-BANTEN": ["Kota Jakarta Selatan", "Kota Jakarta Pusat", "Kota Tangerang", "Kota Tangerang Selatan", "Kota Serang", "Kabupaten Tangerang", "Kota Cilegon", "Kota Jakarta Barat"],
    "EASTERN JABOTABEK": ["Kota Bekasi", "Kabupaten Bekasi", "Kota Depok", "Kota Bogor", "Kabupaten Bogor", "Kota Jakarta Timur", "Kabupaten Karawang", "Kota Jakarta Utara"],
    "JABAR": ["Kota Bandung", "Kota Cirebon", "Kabupaten Bandung", "Kota Sukabumi", "Kota Tasikmalaya", "Kabupaten Garut", "Kota Cimahi", "Kabupaten Sumedang"],
    "JATENG-DIY": ["Kota Semarang", "Kota Yogyakarta", "Kota Surakarta", "Kabupaten Sleman", "Kota Magelang", "Kota Tegal", "Kabupaten Banyumas", "Kota Pekalongan"],
    "JATIM": ["Kota Surabaya", "Kota Malang", "Kota Kediri", "Kabupaten Sidoarjo", "Kota Madiun", "Kabupaten Jember", "Kota Mojokerto", "Kabupaten Banyuwangi"],
    "BALINUSRA": ["Kota Denpasar", "Kabupaten Badung", "Kota Mataram", "Kota Kupang", "Kabupaten Sumbawa", "Kabupaten Buleleng", "Kabupaten Lombok Barat", "Kabupaten Ende"],
    "KALIMANTAN": ["Kota Balikpapan", "Kota Samarinda", "Kota Banjarmasin", "Kota Pontianak", "Kota Palangka Raya", "Kabupaten Kutai Kartanegara", "Kota Tarakan", "Kota Banjarbaru"],
    "SULAWESI": ["Kota Makassar", "Kota Manado", "Kota Palu", "Kota Kendari", "Kota Gorontalo", "Kabupaten Bone", "Kota Pare-Pare", "Kabupaten Minahasa"],
    "PUMA": ["Kota Jayapura", "Kota Ambon", "Kabupaten Merauke", "Kota Sorong", "Kabupaten Mimika", "Kota Ternate", "Kabupaten Jayawijaya", "Kota Manokwari"],
}
CITY_NAME_OVERRIDES = {
    "Kota Pematangsiantar": "KOTA PEMATANG SIANTAR",
    "Kota Tanjungpinang": "KOTA TANJUNG PINANG",
    "Kota Jakarta Selatan": "JAKARTA SELATAN",
    "Kota Jakarta Pusat": "JAKARTA PUSAT",
    "Kota Jakarta Barat": "JAKARTA BARAT",
    "Kota Jakarta Timur": "JAKARTA TIMUR",
    "Kota Jakarta Utara": "JAKARTA UTARA",
    "Kota Palangka Raya": "KOTA PALANGKARAYA",
    "Kota Banjarbaru": "KOTA BANJAR BARU",
    "Kota Manokwari": "MANOKWARI",
    "Kota Sukabumi": "SUKABUMI",  # RETINA-specific: no "KOTA" prefix for this one, unlike the META export
}


def city_xlsx_key(dashboard_name):
    if dashboard_name in CITY_NAME_OVERRIDES:
        return CITY_NAME_OVERRIDES[dashboard_name]
    if dashboard_name.startswith("Kota "):
        return "KOTA " + dashboard_name[len("Kota "):].upper()
    if dashboard_name.startswith("Kabupaten "):
        return dashboard_name[len("Kabupaten "):].upper()
    return dashboard_name.upper()


CITY_XLSX_TO_DASHBOARD = {
    city_xlsx_key(city): (region, city)
    for region, cities in CITY_MAP.items()
    for city in cities
}

# raw RETINA "Operator" values -> the 3 tracked operator groups
OPERATOR_GROUPS = {
    "telkomsel": ["telkomsel", "byu"],
    "ioh": ["indosat", "tri"],
    "xlsmart": ["xl", "axis", "smartfren"],
}
BRAND_TO_GROUP = {brand: group for group, brands in OPERATOR_GROUPS.items() for brand in brands}


def find_data_sheet(wb):
    for name in wb.sheetnames:
        if name != "Sheet1":
            return wb[name]
    raise SystemExit(f"Couldn't find the data sheet among {wb.sheetnames!r} (expected something other than 'Sheet1')")


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = find_data_sheet(wb)
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: i for i, name in enumerate(header)}
    required = ["Store City", "Store Region", "Store Area", "Operator", "Availability Score", "Visibility Score"]
    missing = [c for c in required if c not in col]
    if missing:
        raise SystemExit(f"Missing expected column(s) in {xlsx_path}: {missing} — header was {header!r}")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        operator = r[col["Operator"]]
        if operator is None:
            continue
        avail, vis = r[col["Availability Score"]], r[col["Visibility Score"]]
        if avail is None or vis is None:
            continue
        rows.append({
            "area": r[col["Store Area"]],
            "region": r[col["Store Region"]],
            "city": r[col["Store City"]],
            "branch": r[col.get("Store Branch", -1)] if "Store Branch" in col else None,
            "operator": str(operator).strip().lower(),
            "availability": float(avail),
            "visibility": float(vis),
        })
    return rows


class Accumulator:
    """Running sum+count per (geo key, operator group), so multiple raw
    brands and multiple visit rows all fold into one mean per group."""

    def __init__(self):
        self.sums = defaultdict(lambda: {"availability": 0.0, "visibility": 0.0, "count": 0})
        self.geo_keys = set()

    def add(self, geo_key, group, availability, visibility):
        self.geo_keys.add(geo_key)
        bucket = self.sums[(geo_key, group)]
        bucket["availability"] += availability
        bucket["visibility"] += visibility
        bucket["count"] += 1

    def scores_for(self, geo_key):
        out = {}
        for group in OPERATOR_GROUPS:
            bucket = self.sums.get((geo_key, group))
            if not bucket or bucket["count"] == 0:
                continue
            avail = round(bucket["availability"] / bucket["count"], 2)
            vis = round(bucket["visibility"] / bucket["count"], 2)
            out[group] = {
                "availability": avail,
                "visibility": vis,
                "avIndex": round((avail + vis) / 2, 2),
            }
        return out if out else None


def build(args):
    raw_rows = read_rows(args.xlsx)
    print(f"read {len(raw_rows)} outlet-operator observations from {args.xlsx}", file=sys.stderr)

    acc = Accumulator()
    unmapped_regions, unmapped_cities = set(), set()

    for row in raw_rows:
        group = BRAND_TO_GROUP.get(row["operator"])
        if group is None:
            continue  # unknown brand — shouldn't happen, but don't crash the whole run over it

        region_raw = (row["region"] or "").strip()
        region_key = REGION_XLSX_TO_KEY.get(region_raw.lower())
        if not region_key:
            unmapped_regions.add(region_raw)
            continue
        area_key = REGION_TO_AREA[region_key]

        acc.add(("Nationwide", "ALL", "ALL", "ALL"), group, row["availability"], row["visibility"])
        acc.add(("Area", area_key, "ALL", "ALL"), group, row["availability"], row["visibility"])
        acc.add(("Region", area_key, region_key, "ALL"), group, row["availability"], row["visibility"])

        city_raw = (row["city"] or "").strip().upper()
        match = CITY_XLSX_TO_DASHBOARD.get(city_raw)
        if not match:
            unmapped_cities.add(row["city"])
            continue
        matched_region_key, dashboard_city = match
        acc.add(("City", REGION_TO_AREA[matched_region_key], matched_region_key, dashboard_city), group, row["availability"], row["visibility"])

    if unmapped_regions:
        print(f"warning: {len(unmapped_regions)} unmapped region name(s), skipped: {sorted(unmapped_regions)}", file=sys.stderr)
    if unmapped_cities:
        print(f"note: {len(unmapped_cities)} distinct city name(s) aren't in this dashboard's curated "
              f"CITY_MAP and were skipped from City-level rows (still counted at Region/Area/Nationwide level).",
              file=sys.stderr)

    rows = []
    for level, area, region, city in acc.geo_keys:
        scores = acc.scores_for((level, area, region, city))
        if not scores:
            continue
        rows.append({"level": level, "area": area, "region": region, "city": city, **scores})

    snapshot = {
        "generatedAt": args.month + "-01T00:00:00+07:00",
        "month": args.month,
        "note": "Generated from a RETINA monthly outlet-visit export via scripts/build_retina_snapshot.py — do not hand-edit.",
        "rows": rows,
    }
    return snapshot


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help='path to "Retina - <Month> <Year> (Monthly).xlsx"')
    ap.add_argument("--month", required=True, help="YYYY-MM for this export, e.g. 2026-06")
    ap.add_argument("--out", required=True, help="output snapshot JSON path (dated archive copy)")
    ap.add_argument("--latest-copy", default="retina_latest.json",
                    help="also write here — the stable filename index.html auto-fetches (default: retina_latest.json; pass '' to skip)")
    args = ap.parse_args()

    snapshot = build(args)
    with open(args.out, "w") as f:
        json.dump(snapshot, f, indent=2)
    print(f"wrote {args.out}: {len(snapshot['rows'])} rows")

    if args.latest_copy:
        with open(args.latest_copy, "w") as f:
            json.dump(snapshot, f, indent=2)
        print(f"wrote {args.latest_copy} (auto-fetched by index.html)")


if __name__ == "__main__":
    main()
