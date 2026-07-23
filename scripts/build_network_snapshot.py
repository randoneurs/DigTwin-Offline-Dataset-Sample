#!/usr/bin/env python3
"""Convert an OpenSignal ONX Spotlight export ("ONXSpotlight_YYYYMMDD.xlsx") into
the JSON snapshot index.html's Network Experience Benchmark card loads.

Usage:
  python3 scripts/build_network_snapshot.py \
      --xlsx ONXSpotlight_20260701.xlsx \
      --out network_20260701.json

Source: one sheet, one row per (Location, Operator) combination, ~550 columns.
Relevant columns: End_Date (the reporting period's reference date), Location (a
single flat field mixing every grain — "NATIONWIDE", "AREA 1".."AREA 4", the 12
region names, or a city/kabupaten name — there's no separate level column, so
level is inferred by membership in known Nationwide/Area/Region name sets, with
everything else treated as City), Device_SIMServiceProviderBrandName (operator;
note "Three" is exported as the bare number 3 rather than text — likely the "3"
brand mark getting numeric-coerced during export), and MEAN_<Parameter>_Overall
/ PERCENT_<Parameter>_Overall for each of the 12 tracked parameters (CCQ, ECQ,
and ConsistentQuality only ever have a PERCENT column, never MEAN, unlike the
other 9 — this is a real gap against "use the column with MEAN in the title"
for those three, not a bug, so PERCENT is substituted for just those three).

Aggregation: average every raw-brand Overall value at each geography level
(City, Region, Area, Nationwide) — a plain mean, no weighting — then combine
Indosat+Three into "ioh" and XL+Smartfren into "xlsmart" (also a plain mean);
Telkomsel has no second brand in this export, so its group is just itself.
"""
import argparse
import json
import sys
from collections import defaultdict

import openpyxl

AREA_MAP = {
    "Area 1": ["SUMBAGUT", "SUMBAGTENG", "SUMBAGSEL"],
    "Area 2": ["JAKARTA-BANTEN", "EASTERN JABOTABEK", "JABAR"],
    "Area 3": ["JATENG-DIY", "JATIM", "BALINUSRA"],
    "Area 4": ["KALIMANTAN", "SULAWESI", "PUMA"],
}
REGION_TO_AREA = {region: area for area, regions in AREA_MAP.items() for region in regions}

# ONX's "Location" column (any case) -> this dashboard's REGION_LABELS key. Note
# JAKARTA-BANTEN never appears as its own Location row in this export (only the
# 5 constituent Jakarta cities do) — its Region-level row will simply be absent,
# same graceful-gap handling as everything else here.
REGION_LOCATION_TO_KEY = {
    "sumbagut": "SUMBAGUT",
    "sumbagteng": "SUMBAGTENG",
    "sumbagsel": "SUMBAGSEL",
    "jakarta banten": "JAKARTA-BANTEN",
    "eastern jabotabek": "EASTERN JABOTABEK",
    "jabar": "JABAR",
    "jateng-diy": "JATENG-DIY",
    "jatim": "JATIM",
    "bali nusra": "BALINUSRA",
    "kalimantan": "KALIMANTAN",
    "sulawesi": "SULAWESI",
    "maluku dan papua": "PUMA",
}
AREA_LOCATIONS = {"area 1": "Area 1", "area 2": "Area 2", "area 3": "Area 3", "area 4": "Area 4"}
NATIONWIDE_LOCATION = "nationwide"

# The dashboard's curated CITY_MAP (index.html) — only these cities resolve
# through the fArea/fRegion/fCity filters.
CITY_MAP = {
    "SUMBAGUT": ["Kota Medan", "Kota Banda Aceh", "Kabupaten Deli Serdang", "Kota Pematangsiantar", "Kota Binjai", "Kabupaten Langkat", "Kota Lhokseumawe", "Kabupaten Simalungun"],
    "SUMBAGTENG": ["Kota Pekanbaru", "Kota Padang", "Kota Batam", "Kota Jambi", "Kota Dumai", "Kabupaten Kampar", "Kota Tanjungpinang", "Kabupaten Bungo"],
    "SUMBAGSEL": ["Kota Palembang", "Kota Bandar Lampung", "Kota Bengkulu", "Kota Pangkalpinang", "Kabupaten Ogan Ilir", "Kota Prabumulih", "Kabupaten Lampung Selatan", "Kota Metro"],
    "JAKARTA-BANTEN": ["Kota Jakarta Selatan", "Kota Jakarta Pusat", "Kota Tangerang", "Kota Tangerang Selatan", "Kota Serang", "Kabupaten Tangerang", "Kota Cilegon", "Kota Jakarta Barat"],
    "EASTERN JABOTABEK": ["Kota Bekasi", "Kabupaten Bekasi", "Kota Depok", "Kota Bogor", "Kabupaten Bogor", "Kota Jakarta Timur", "Kabupaten Karawang", "Kota Jakarta Utara"],
    "JABAR": ["Kota Bandung", "Kota Cirebon", "Kabupaten Bandung", "Kota Sukabumi", "Kota Tasikmalaya", "Kabupaten Garut", "Kota Cimahi", "Kabupaten Sumedang"],
    "JATENG-DIY": ["Kota Semarang", "Kota Yogyakarta", "Kota Surakarta", "Kabupaten Sleman", "Kota Magelang", "Kota Tegal", "Kabupaten Banyumas", "Kota Pekalongan"],
    "JATIM": ["Kota Surabaya", "Kota Malang", "Kota Kediri", "Kabupaten Sidoarjo", "Kota Madiun", "Kabupaten Jember", "Kota Mojokerto", "Kabupaten Banyuwangi"],
    "BALINUSRA": ["Kota Denpasar", "Kabupaten Badung", "Kota Mataram", "Kota Kupang", "Kabupaten Sumbawa", "Kabupaten Buleleng", "Kabupaten Lombok Barat", "Kabupaten Ende"],
    "KALIMANTAN": ["Kota Balikpapan", "Kota Samarinda", "Kota Banjarmasin", "Kota Pontianak", "Kota Palangka Raya", "Kabupaten Kutai Kartanegara", "Kota Tarakan", "Kota Banjarbaru"],
    "SULAWESI": ["Kota Makassar", "Kota Manado", "Kota Palu", "Kota Kendari", "Kota Gorontalo", "Kabupaten Bone", "Kota Pare-Pare", "Kabupaten Minahasa"],
    "PUMA": ["Kota Jayapura", "Kota Ambon", "Kabupaten Merauke", "Kota Sorong", "Kabupaten Mimika", "Kota Ternate", "Kabupaten Jayawijaya", "Kota Manokwari"],
}
# ONX-specific spelling quirks — checked against a real export; deliberately NOT
# reused from the RETINA/META overrides because several of these differ (e.g.
# Jakarta cities and Sukabumi keep their "KOTA " prefix here, unlike RETINA).
CITY_NAME_OVERRIDES = {
    "Kota Pematangsiantar": "KOTA PEMATANG SIANTAR",
    "Kota Tanjungpinang": "KOTA TANJUNG PINANG",
    "Kota Palangka Raya": "KOTA PALANGKARAYA",
    "Kota Banjarbaru": "KOTA BANJAR BARU",
    "Kota Manokwari": "MANOKWARI",
    "Kota Pangkalpinang": "KOTA PANGKAL PINANG",
    "Kota Pare-Pare": "KOTA PAREPARE",
}


def city_location_key(dashboard_name):
    if dashboard_name in CITY_NAME_OVERRIDES:
        return CITY_NAME_OVERRIDES[dashboard_name]
    if dashboard_name.startswith("Kota "):
        return "KOTA " + dashboard_name[len("Kota "):].upper()
    if dashboard_name.startswith("Kabupaten "):
        return dashboard_name[len("Kabupaten "):].upper()
    return dashboard_name.upper()


CITY_LOCATION_TO_DASHBOARD = {
    city_location_key(city): (region, city)
    for region, cities in CITY_MAP.items()
    for city in cities
}

# ONX raw brand -> operator group. "Three" is exported as the number 3.
BRAND_TO_GROUP = {
    "telkomsel": "telkomsel",
    "indosat": "ioh",
    "3": "ioh",       # "Three" / Tri, numeric-coerced during export
    "xl": "xlsmart",
    "smartfren": "xlsmart",
}

# label -> source column. CCQ/ECQ/ConsistentQuality only ever have a PERCENT_
# column (no MEAN_ variant exists for them in this export).
PARAMETERS = [
    {"key": "downloadSpeed", "label": "DownloadSpeed (Mbps)", "column": "MEAN_DownloadSpeed_Overall"},
    {"key": "uploadSpeed", "label": "UploadSpeed (Mbps)", "column": "MEAN_UploadSpeed_Overall"},
    {"key": "videoExperience", "label": "VideoExperience", "column": "MEAN_VideoExperience_Overall"},
    {"key": "voiceAppExperience", "label": "VoiceAppExperience", "column": "MEAN_VoiceAppExperience_Overall"},
    {"key": "gamesExperience", "label": "GamesExperience", "column": "MEAN_GamesExperience_Overall"},
    {"key": "availability", "label": "Availability", "column": "MEAN_Availability_AllUser_Overall"},
    {"key": "ccq", "label": "CCQ : Core Consistent Quality", "column": "PERCENT_CCQ_Overall"},
    {"key": "ecq", "label": "ECQ : Excellent Consistent Quality", "column": "PERCENT_ECQ_Overall"},
    {"key": "liveVideoExperience", "label": "LiveVideoExperience", "column": "MEAN_LiveVideoExperience_Overall"},
    {"key": "consistentQuality", "label": "ConsistentQuality", "column": "PERCENT_ConsistentQuality_Overall"},
    {"key": "coverageExperience", "label": "CoverageExperience", "column": "MEAN_CoverageExperience_Overall"},
    {"key": "reliability", "label": "Reliability", "column": "MEAN_Reliability_Overall"},
]
DEFAULT_PARAM_KEY = "videoExperience"


def classify_location(raw):
    loc = raw.strip()
    lower = loc.lower()
    if lower == NATIONWIDE_LOCATION:
        return ("Nationwide", "ALL", "ALL", "ALL")
    if lower in AREA_LOCATIONS:
        return ("Area", AREA_LOCATIONS[lower], "ALL", "ALL")
    if lower in REGION_LOCATION_TO_KEY:
        region_key = REGION_LOCATION_TO_KEY[lower]
        return ("Region", REGION_TO_AREA[region_key], region_key, "ALL")
    match = CITY_LOCATION_TO_DASHBOARD.get(loc.upper())
    if match:
        region_key, dashboard_city = match
        return ("City", REGION_TO_AREA[region_key], region_key, dashboard_city)
    return None  # a real kabupaten/kota this dashboard doesn't have a curated filter for


def find_data_sheet(wb):
    for name in wb.sheetnames:
        return wb[name]
    raise SystemExit("workbook has no sheets")


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = find_data_sheet(wb)
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: i for i, name in enumerate(header)}
    required = ["End_Date", "Location", "Device_SIMServiceProviderBrandName"] + [p["column"] for p in PARAMETERS]
    missing = [c for c in required if c not in col]
    if missing:
        raise SystemExit(f"Missing expected column(s) in {xlsx_path}: {missing}")

    end_dates = set()
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        brand_raw = r[col["Device_SIMServiceProviderBrandName"]]
        if brand_raw is None:
            continue
        group = BRAND_TO_GROUP.get(str(brand_raw).strip().lower())
        if group is None:
            continue
        end_dates.add(r[col["End_Date"]])
        values = {}
        for p in PARAMETERS:
            v = r[col[p["column"]]]
            if v is not None:
                values[p["key"]] = float(v)
        rows.append({"location": r[col["Location"]], "group": group, "values": values})
    return rows, end_dates


class Accumulator:
    def __init__(self):
        self.sums = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))  # geo_key -> group -> [sum, count]
        self.geo_keys = set()

    def add_values(self, geo_key, group, values):
        self.geo_keys.add(geo_key)
        store = self.sums[(geo_key, group)]
        for key, val in values.items():
            entry = store[key]
            entry[0] += val
            entry[1] += 1

    def scores_for(self, geo_key):
        out = {}
        for group in ("telkomsel", "ioh", "xlsmart"):
            store = self.sums.get((geo_key, group))
            if not store:
                continue
            group_scores = {}
            for p in PARAMETERS:
                entry = store.get(p["key"])
                if entry and entry[1] > 0:
                    group_scores[p["key"]] = round(entry[0] / entry[1], 2)
            if group_scores:
                out[group] = group_scores
        return out if out else None


def build(args):
    raw_rows, end_dates = read_rows(args.xlsx)
    print(f"read {len(raw_rows)} location-operator observations from {args.xlsx}", file=sys.stderr)
    if len(end_dates) > 1:
        print(f"warning: multiple End_Date values found: {end_dates}, using the latest", file=sys.stderr)
    end_date = sorted(str(d) for d in end_dates)[-1] if end_dates else args.as_of
    as_of = args.as_of or str(end_date).split(" ")[0]

    acc = Accumulator()
    unclassified = set()

    for row in raw_rows:
        classified = classify_location(row["location"])
        if not classified:
            unclassified.add(row["location"])
            continue
        level, area, region, city = classified
        acc.add_values((level, area, region, city), row["group"], row["values"])

    if unclassified:
        print(f"note: {len(unclassified)} Location value(s) aren't Nationwide/Area/Region and aren't in this "
              f"dashboard's curated CITY_MAP — skipped (real kabupaten/kota outside the curated filter set, not an error).",
              file=sys.stderr)

    rows = []
    for level, area, region, city in acc.geo_keys:
        scores = acc.scores_for((level, area, region, city))
        if not scores:
            continue
        rows.append({"level": level, "area": area, "region": region, "city": city, **scores})

    return {
        "generatedAt": as_of + "T00:00:00+07:00",
        "asOfDate": as_of,
        "note": "Generated from an OpenSignal ONX Spotlight export via scripts/build_network_snapshot.py — do not hand-edit.",
        "parameters": [{"key": p["key"], "label": p["label"]} for p in PARAMETERS],
        "defaultParam": DEFAULT_PARAM_KEY,
        "rows": rows,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="path to ONXSpotlight_YYYYMMDD.xlsx")
    ap.add_argument("--as-of", default=None, help="override the as-of date (YYYY-MM-DD); default: the export's own End_Date")
    ap.add_argument("--out", required=True, help="output snapshot JSON path (dated archive copy)")
    ap.add_argument("--latest-copy", default="network_latest.json",
                    help="also write here — the stable filename index.html auto-fetches (default: network_latest.json; pass '' to skip)")
    args = ap.parse_args()

    snapshot = build(args)
    with open(args.out, "w") as f:
        json.dump(snapshot, f, indent=2)
    print(f"wrote {args.out}: {len(snapshot['rows'])} rows")

    if args.latest_copy:
        with open(args.latest_copy, "w") as f:
            json.dump(snapshot, f, indent=2)
        print(f"wrote {args.latest_copy} (auto-fetched by index.html)")


if __name__ == "__main__":
    main()
